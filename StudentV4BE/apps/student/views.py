"""
File Name: views.py
Author: LiuHao
Date: 2021-10-24
"""
from django.http import JsonResponse
from django.db.models import Q
from django.conf import settings

from .models import Student

import os
import json
import uuid
import hashlib
import openpyxl
# Create your views here.


def get_student(request):

    """ 获取所有学生的信息 """

    try:
        # 使用ORM获取所有学生的信息
        obj_student = Student.objects.all().values()
        # 把结果转为list
        student = list(obj_student)

        return JsonResponse({
            'code': 1,
            'data': student,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def query_students(request):

    """ 查询学生信息 """

    data = json.loads(request.body.decode('utf-8'))
    try:
        # 使用ORM获取所有学生的信息
        obj_student = Student.objects.filter(
            Q(sno__icontains=data['input_str'])    | Q(name__icontains=data['input_str']) |
            Q(gender__icontains=data['input_str']) | Q(mobile__icontains=data['input_str']) |
            Q(email__icontains=data['input_str'])  | Q(address__icontains=data['input_str'])
        ).values()
        # 把结果转为list
        student = list(obj_student)

        return JsonResponse({
            'code': 1,
            'data': student,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def is_exist_sno(request):

    """ 判断学号是否存在 """

    data = json.loads(request.body.decode('utf-8'))
    obj_student = Student.objects.filter(sno=data['sno'])

    try:
        if obj_student.count() == 0:
            return JsonResponse({
                'code': 1,
                'exists': False
            })
        else:
            return JsonResponse({
                'code': 1,
                'exists': True
            })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def add_student(request):
    """ 添加学生到数据库中 """
    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_student = Student(
            sno=data['sno'], name=data['name'], gender=data['gender'], birthday=data['birthday'],
            mobile=data['mobile'], email=data['email'], address=data['address'],
            image=data['image']
        )
        obj_student.save()
        # 使用ORM获取所有学生的信息
        obj_student = Student.objects.all().values()
        # 把结果转为list
        student = list(obj_student)
        return JsonResponse({
            'code': 1,
            'data': student,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def update_student(request):

    """ 更新数据到数据库 """

    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.name     = data['name']
        obj_student.mobile   = data['mobile']
        obj_student.birthday = data['birthday']
        obj_student.gender   = data['gender']
        obj_student.email    = data['email']
        obj_student.address  = data['address']
        obj_student.image = data['image']

        obj_student.save()

        obj_students = Student.objects.all().values()

        students = list(obj_students)
        return JsonResponse({
            'code': 1,
            'data': students,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def delete_student(request):

    """ 删除一条消息 """

    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.delete()
        obj_students = Student.objects.all().values()

        students = list(obj_students)
        return JsonResponse({
            'code': 1,
            'data': students,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def delete_students(request):

    """ 删除多条数据 """

    data = json.loads(request.body.decode('utf-8'))
    try:
        for one_student in data['students']:
            obj_student = Student.objects.get(sno=one_student['sno'])
            obj_student.delete()
        obj_students = Student.objects.all().values()

        students = list(obj_students)
        return JsonResponse({
            'code': 1,
            'data': students,
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def get_random_str():
    # 获取uuid的随机数
    uuid_val = uuid.uuid4()
    # 获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    # 获取md5实例
    md5 = hashlib.md5()
    # 拿取uuid的md5摘要
    md5.update(uuid_str)
    # 返回固定长度的字符串
    return md5.hexdigest()


def upload(request):

    """ 上传图片 """

    rev_file = request.FILES.get('avatar')
    # 判断是否有文件
    if not rev_file:
        return JsonResponse({
            'code': 0,
            'msg': "no exist"
        })
    # 获得唯一的名字：uuid+hash
    new_name = get_random_str()
    # 准备写入的url
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])

    # 开始写入到本地磁盘
    try:
        f = open(file_path, 'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        f.close()
        return JsonResponse({
            'code': 1,
            'name': new_name + os.path.splitext(rev_file.name)[1],
        })
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })


def read_excel_dict(path: str):
    """ 读取excel数据 存储为字典 """
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['student']
    # 定义一个变量存储最终的数据
    students = list()
    # 准备key
    keys = [
        'sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address',
    ]
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = dict()
        for index, cell in enumerate(row):
            temp_dict[keys[index]] = cell.value
        students.append(temp_dict)

    return students


def import_student_excel(request):

    """ 从excel批量导入学生信息 """

    # 1.接收excel文件到media文件夹
    rev_file = request.FILES.get('excel')
    # 判断是否有文件
    if not rev_file:
        return JsonResponse({
            'code': 0,
            'msg': "file is not exist"
        })
    # 获得唯一的名字：uuid+hash
    new_name = get_random_str()
    # 准备写入的url
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path, 'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        f.close()
    except Exception as e:
        return JsonResponse({
            'code': 0,
            'msg': 'error{}'.format(str(e)),
        })

    # 2.读取存储在media文件夹的数据
    ex_students = read_excel_dict(file_path)

    # 3.把读取的数据存储到数据库
    success = 0
    error = 0
    error_sno = list()

    for one_student in ex_students:
        # noinspection PyBroadException
        try:
            # 如果sno一样则不符合同一性
            Student.objects.create(
                sno=one_student['sno'], name=one_student['name'], gender=one_student['gender'],
                birthday=one_student['birthday'], mobile=one_student['mobile'],
                email=one_student['email'], address=one_student['address'],
            )
            # obj_student.save()

            success += 1
        except:
            error += 1
            error_sno.append(one_student['sno'])

    # 4.返回---导入的信息
    obj_students = Student.objects.all().values()
    students = list(obj_students)
    return JsonResponse({
        'code': 1,
        'success': success,
        'error': error,
        'errors': error_sno,
        'data': students,
    })


def write_to_excel(data: list, path: str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'student'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)


def export_student_excel(request):

    """ 到处数据到excel """

    obj_students = Student.objects.all().values()
    students = list(obj_students)

    # 准备名称
    excel_name = get_random_str() + '.xlsx'
    # 准备写入的路径
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到excel
    write_to_excel(students, path)

    return JsonResponse({
        'code': 1,
        'name': excel_name,
    })
